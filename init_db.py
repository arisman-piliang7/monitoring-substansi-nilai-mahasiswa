"""
Jalankan sekali untuk membuat exam.db dari file Excel export di folder data/,
lengkap dengan tabel users (admin + mahasiswa), durasi pertemuan online,
Bintang, dan Diskusi Mteam.

Jalankan ulang skrip ini akan MEREBUILD ulang tabel `students` dan
`meeting_durations` (nilai selalu sinkron dengan Excel) tapi TIDAK
menghapus password yang sudah dibuat mahasiswa di tabel `users`
(aman untuk re-run).

Penggunaan:
    python init_db.py
"""
import re
import sqlite3
from pathlib import Path

import openpyxl

from auth import get_conn, hash_password, now_iso, DB_PATH

DATA_DIR = Path(__file__).parent / "data"

XLSX_FILES = {
    "A": DATA_DIR / "IF_A_SORE.xlsx",
    "C": DATA_DIR / "IF_C_SORE.xlsx",
}

# Kolom-kolom pada baris export (1-indexed) yang posisinya TETAP di semua file.
COL_NIM = 2
COL_NAMA = 3
COL_TUGAS = 6
COL_UTS = 7
COL_MEETING_START = 8   # H
COL_MEETING_END = 14    # N (dinamis - bisa None jika kelas punya lebih sedikit meeting)
COL_ABSENSI = 15        # O
COL_BINTANG = 16        # P
COL_DISKUSI = 17        # Q
COL_FORMATIF = 19       # S
COL_TUGAS2 = 20         # T
COL_UAS = 21            # U
COL_NILAI_AKTUAL = 23   # W
COL_NILAI_HURUF = 24    # X

ROW_TARGET_DURASI = 9   # H9:N9 = target/acuan durasi per meeting
ROW_HEADER = 10         # H10:N10 = kode meeting (M03, M05, ...)
ROW_DATA_START = 11

# Ganti ini setelah first-run demi keamanan (lihat pesan di akhir skrip)
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "ubah-password-ini"


SCHEMA = """
CREATE TABLE IF NOT EXISTS students (
    nim           TEXT PRIMARY KEY,
    nama          TEXT NOT NULL,
    kelas         TEXT NOT NULL,
    tugas         REAL,
    uts           REAL,
    formatif      REAL,
    tugas2        REAL,
    uas           REAL,
    bintang       REAL,
    diskusi_mteam REAL,
    nilai_aktual  REAL,
    nilai_huruf   TEXT,
    absensi       REAL
);

CREATE TABLE IF NOT EXISTS meeting_durations (
    nim               TEXT NOT NULL,
    kelas             TEXT NOT NULL,
    meeting_code      TEXT NOT NULL,
    urutan            INTEGER NOT NULL,
    duration_seconds  INTEGER,
    duration_display  TEXT,
    target_seconds    INTEGER,
    target_display    TEXT,
    PRIMARY KEY (nim, meeting_code),
    FOREIGN KEY (nim) REFERENCES students (nim)
);

CREATE TABLE IF NOT EXISTS users (
    username        TEXT PRIMARY KEY,
    role            TEXT NOT NULL CHECK (role IN ('admin', 'student')),
    nim             TEXT,
    password_hash   TEXT,
    salt            TEXT,
    is_first_login  INTEGER NOT NULL DEFAULT 1,
    created_at      TEXT,
    updated_at      TEXT,
    FOREIGN KEY (nim) REFERENCES students (nim)
);

CREATE TABLE IF NOT EXISTS reset_requests (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    nim           TEXT NOT NULL,
    requested_at  TEXT NOT NULL,
    status        TEXT NOT NULL DEFAULT 'pending' CHECK (status IN ('pending', 'approved', 'rejected')),
    resolved_at   TEXT,
    FOREIGN KEY (nim) REFERENCES students (nim)
);
"""


def build_schema(conn: sqlite3.Connection):
    conn.executescript(SCHEMA)
    conn.commit()


# ----------------------------------------------------------------------------
# Duration parsing helpers ("1h 24m 16s", "49m 2s", "23m 10s" -> detik)
# ----------------------------------------------------------------------------
_DURATION_RE = re.compile(
    r"(?:(?P<h>\d+)h)?\s*(?:(?P<m>\d+)m)?\s*(?:(?P<s>\d+)s)?"
)


def parse_duration(text) -> int | None:
    """Mengubah string durasi seperti '1h 24m 16s' menjadi total detik."""
    if text is None:
        return None
    text = str(text).strip()
    if not text:
        return None
    match = _DURATION_RE.fullmatch(text)
    if not match or not any(match.groups()):
        return None
    h = int(match.group("h") or 0)
    m = int(match.group("m") or 0)
    s = int(match.group("s") or 0)
    return h * 3600 + m * 60 + s


def _num(value):
    """Konversi aman ke float, mengembalikan None untuk sel kosong."""
    if value is None or value == "":
        return None
    return float(value)


# ----------------------------------------------------------------------------
# Load students + meeting durations directly from the Excel exports
# ----------------------------------------------------------------------------
def load_students(conn: sqlite3.Connection):
    conn.execute("DELETE FROM students")
    conn.execute("DELETE FROM meeting_durations")

    for kelas, path in XLSX_FILES.items():
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        # Kode meeting & target durasi per kelas (kolom H..N, posisi tetap)
        meeting_cols = []  # list of (col_index, urutan, meeting_code, target_seconds, target_display)
        for urutan, col in enumerate(range(COL_MEETING_START, COL_MEETING_END + 1), start=1):
            code = ws.cell(row=ROW_HEADER, column=col).value
            if code is None or str(code).strip() == "":
                continue
            target_display = ws.cell(row=ROW_TARGET_DURASI, column=col).value
            meeting_cols.append((
                col, urutan, str(code).strip(),
                parse_duration(target_display), target_display,
            ))

        student_rows = []
        meeting_rows = []

        for r in range(ROW_DATA_START, ws.max_row + 1):
            nim_val = ws.cell(row=r, column=COL_NIM).value
            if nim_val is None or str(nim_val).strip() == "":
                continue  # baris kosong / pemisah
            nim = str(nim_val).strip()
            nama = str(ws.cell(row=r, column=COL_NAMA).value or "").strip()

            absensi_raw = _num(ws.cell(row=r, column=COL_ABSENSI).value)
            nilai_huruf_raw = ws.cell(row=r, column=COL_NILAI_HURUF).value

            student_rows.append((
                nim,
                nama,
                kelas,
                _num(ws.cell(row=r, column=COL_TUGAS).value),
                _num(ws.cell(row=r, column=COL_UTS).value),
                _num(ws.cell(row=r, column=COL_FORMATIF).value),
                _num(ws.cell(row=r, column=COL_TUGAS2).value),
                _num(ws.cell(row=r, column=COL_UAS).value),
                _num(ws.cell(row=r, column=COL_BINTANG).value),
                _num(ws.cell(row=r, column=COL_DISKUSI).value),
                _num(ws.cell(row=r, column=COL_NILAI_AKTUAL).value),
                str(nilai_huruf_raw).strip() if nilai_huruf_raw else None,
                round(absensi_raw * 100, 0) if absensi_raw is not None else None,
            ))

            for col, urutan, code, target_seconds, target_display in meeting_cols:
                dur_display = ws.cell(row=r, column=col).value
                meeting_rows.append((
                    nim, kelas, code, urutan,
                    parse_duration(dur_display), dur_display,
                    target_seconds, target_display,
                ))

        conn.executemany(
            """INSERT INTO students
               (nim, nama, kelas, tugas, uts, formatif, tugas2, uas,
                bintang, diskusi_mteam, nilai_aktual, nilai_huruf, absensi)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            student_rows,
        )
        conn.executemany(
            """INSERT INTO meeting_durations
               (nim, kelas, meeting_code, urutan, duration_seconds, duration_display,
                target_seconds, target_display)
               VALUES (?,?,?,?,?,?,?,?)""",
            meeting_rows,
        )
    conn.commit()


def sync_student_users(conn: sqlite3.Connection):
    """Create a 'student' user row for every student that doesn't have one yet.
    Existing users (and their passwords) are left untouched."""
    cur = conn.execute("SELECT nim FROM students")
    nims = [row["nim"] for row in cur.fetchall()]

    for nim in nims:
        existing = conn.execute(
            "SELECT 1 FROM users WHERE username = ?", (nim,)
        ).fetchone()
        if existing is None:
            conn.execute(
                """INSERT INTO users
                   (username, role, nim, password_hash, salt, is_first_login, created_at, updated_at)
                   VALUES (?, 'student', ?, NULL, NULL, 1, ?, ?)""",
                (nim, nim, now_iso(), now_iso()),
            )
    conn.commit()


def ensure_admin(conn: sqlite3.Connection):
    existing = conn.execute(
        "SELECT 1 FROM users WHERE username = ?", (DEFAULT_ADMIN_USERNAME,)
    ).fetchone()
    if existing is None:
        hashed, salt = hash_password(DEFAULT_ADMIN_PASSWORD)
        conn.execute(
            """INSERT INTO users
               (username, role, nim, password_hash, salt, is_first_login, created_at, updated_at)
               VALUES (?, 'admin', NULL, ?, ?, 1, ?, ?)""",
            (DEFAULT_ADMIN_USERNAME, hashed, salt, now_iso(), now_iso()),
        )
        conn.commit()
        return True
    return False


def main():
    is_new_db = not DB_PATH.exists()
    conn = get_conn()
    try:
        build_schema(conn)
        load_students(conn)
        sync_student_users(conn)
        admin_created = ensure_admin(conn)
    finally:
        conn.close()

    print(f"Database siap di: {DB_PATH}")
    if admin_created:
        print("\n=== AKUN ADMIN DIBUAT ===")
        print(f"Username : {DEFAULT_ADMIN_USERNAME}")
        print(f"Password : {DEFAULT_ADMIN_PASSWORD}")
        print("PENTING: login sebagai admin lalu segera ganti password ini")
        print("melalui menu 'Ganti Password' di sidebar aplikasi.")
    else:
        print("Akun admin sudah ada sebelumnya (tidak diubah).")
    print("\nMahasiswa akan diminta membuat password sendiri saat login")
    print("pertama kali menggunakan NIM masing-masing.")


if __name__ == "__main__":
    main()
